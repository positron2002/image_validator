import streamlit as st
import pandas as pd
import json
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

# Streamlit page configuration
st.set_page_config(layout="wide")
st.title("A-PAG QC LOG")

# File uploader
uploaded_file = st.file_uploader("Upload your CSV or Excel file", type=["csv", "xlsx"])

# Feedback file location
FEEDBACK_FILE = "feedback_data.json"

# Ensure the feedback file exists
if not os.path.exists(FEEDBACK_FILE):
    with open(FEEDBACK_FILE, "w") as f:
        json.dump({}, f)

# Load previous feedback data into session state
if "feedback" not in st.session_state:
    with open(FEEDBACK_FILE, "r") as f:
        st.session_state.feedback = json.load(f)

# Pagination setup
ROWS_PER_PAGE = 10
if "page" not in st.session_state:
    st.session_state.page = 0  # Start on page 0

# Disapproval reasons
disapproval_reasons = [
    "Wrong Before Image/Poor Identification",
    "After Photo-Missing",
    "After Photo-Wrong/Blurry",
    "Incomplete Work/Work Not Started",
    "Image taken from wrong angle"
]

if uploaded_file:
    try:
        df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith(".csv") else pd.read_excel(uploaded_file)
    except Exception as e:
        st.error(f"Error reading the file: {e}")
        st.stop()

    required_cols = {"Project Id", "Raised Evidence", "Latest Evidence", "Zone", "Ward", "Action Item"}
    if not required_cols.issubset(df.columns):
        st.error(f"Your file is missing required columns: {required_cols - set(df.columns)}")
        st.stop()

    # Filters
    selected_action_item = st.selectbox("Filter by Action Item", ["All"] + df["Action Item"].dropna().unique().tolist())
    selected_zone = st.selectbox("Filter by Zone", ["All"] + df["Zone"].dropna().unique().tolist())
    selected_ward = st.selectbox("Filter by Ward", ["All"] + df["Ward"].dropna().unique().tolist())

    # Apply filters
    filtered_df = df.copy()
    if selected_action_item != "All":
        filtered_df = filtered_df[filtered_df["Action Item"] == selected_action_item]
    if selected_zone != "All":
        filtered_df = filtered_df[filtered_df["Zone"] == selected_zone]
    if selected_ward != "All":
        filtered_df = filtered_df[filtered_df["Ward"] == selected_ward]

    # Reset pagination if filtered data is smaller than the current page
    total_pages = len(filtered_df) // ROWS_PER_PAGE + (len(filtered_df) % ROWS_PER_PAGE > 0)
    if st.session_state.page >= total_pages:
        st.session_state.page = 0

    # Paginate filtered results
    start_idx = st.session_state.page * ROWS_PER_PAGE
    end_idx = start_idx + ROWS_PER_PAGE
    df_page = filtered_df.iloc[start_idx:end_idx]

    # Live Review Summary
    status_counts = {
        "Status Yet to be Updated": 0,
        "Not Reviewed": 0,
        "Correct": 0,
        "Incorrect": 0
    }

    for pid in filtered_df["Project Id"]:
        current_status = st.session_state.feedback.get(str(pid), {}).get("Quality", "Status Yet to be Updated")
        status_counts[current_status] += 1

    # Sidebar Summary
    with st.sidebar:
        st.subheader("Review Summary 📊")
        st.write(f" **Correct:** {status_counts['Correct']}")
        st.write(f"**Incorrect:** {status_counts['Incorrect']}")
        st.write(f" **Not Reviewed:** {status_counts['Not Reviewed']}")
        st.write(f" **Status Yet to be Updated:** {status_counts['Status Yet to be Updated']}")

        # Live Percentage calculation
        total_correct = status_counts['Correct']
        total_incorrect = status_counts['Incorrect']
        total_not_reviewed = status_counts['Not Reviewed']
        total_status_not_updated = status_counts['Status Yet to be Updated']
        if total_correct + total_incorrect > 0:
            live_percentage = (total_correct * 100) / (total_correct + total_incorrect)
            st.write(f" **Current QC Status %:** {live_percentage:.2f}%")

        # Status % Calculation
        total_reviewed = total_correct + total_incorrect
        total_possible = total_reviewed+total_not_reviewed+total_status_not_updated
        
        if total_possible>0:
            status_percentage = (total_reviewed * 100) / total_possible
            st.write(f"**Current Sample Size %** :{status_percentage:.2f}%")
        else:
            st.write("**Current Sample Size %** :0.00%")
        
        #Number of QC done
        st.write(f"**Number of QC Done:** {total_correct + total_incorrect}")
        
    # Check if the page is empty
    if df_page.empty:
        st.error(f"Page {st.session_state.page} is empty. Total filtered rows: {len(filtered_df)}")
    else:
        # Loop through rows
        for _, row in df_page.iterrows():
            project_id = str(row["Project Id"])
            pre_image = row["Raised Evidence"]
            post_image = row["Latest Evidence"]
            zone = row["Zone"]
            ward = row["Ward"]
            latest_comment = row.get("Latest Comment", "No comment available")

            st.subheader(f"Project ID: {project_id} | Action Item: {row.get('Action Item', 'N/A')}")
            st.text(f"Zone: {zone} | Ward: {ward}")
            st.text(f"Raised Comment: {row.get('Raised Comment', 'N/A')}")
            st.text(f"Latest Comment: {latest_comment}")
            
            # Display images
            col1, col2 = st.columns(2)
            with col1:
                if pd.notna(pre_image):
                   st.markdown(
    f'<img src="{pre_image}" style="max-width: 500px; max-height: 500px;">',
    unsafe_allow_html=True
)
                else:
                    st.warning("No Pre Image Provided")

            with col2:
                if pd.notna(post_image):
                    st.markdown(
    f'<img src="{post_image}" style="max-width: 500px; max-height: 500px;">',
    unsafe_allow_html=True
)
                else:
                    st.warning("No Post Image Provided")

            saved_status = st.session_state.feedback.get(project_id, {}).get("Quality", "Status Yet to be Updated")

            status = st.radio(
                label=f"Status for Project ID {project_id}",
                options=["Status Yet to be Updated", "Not Reviewed", "Correct", "Incorrect"],
                key=f"status_{project_id}",
                index=["Status Yet to be Updated", "Not Reviewed", "Correct", "Incorrect"].index(saved_status)
            )

            saved_reason = st.session_state.feedback.get(project_id, {}).get("comment", "")

            reason = ""
            if status == "Incorrect":
                reason = st.selectbox(
                    label=f"Reason for Disapproval (ID {project_id})",
                    options=disapproval_reasons,
                    key=f"reason_{project_id}",
                    index=disapproval_reasons.index(saved_reason) if saved_reason in disapproval_reasons else 0
                )

            # Store feedback in session state
            st.session_state.feedback[project_id] = {
                "Quality": status,
                "comment": reason if status == "Incorrect" else ""
            }

        # Pagination controls
        total_pages = len(filtered_df) // ROWS_PER_PAGE + (len(filtered_df) % ROWS_PER_PAGE > 0)

        st.markdown(f"**Page {st.session_state.page + 1} of {total_pages}**")  # Show current page

        col1, col2, col3 = st.columns([1, 6, 1])

        with col1:
            if st.session_state.page > 0:
                if st.button("⬅️ Previous Page"):
                    st.session_state.page -= 1
                    st.rerun()

        with col2:
            page_numbers = list(range(max(0, st.session_state.page - 3), min(total_pages, st.session_state.page + 10)))
            pages = st.columns(len(page_numbers))
            for i, p in enumerate(page_numbers): 
                with pages[i]:
                    if st.button(f"**{p + 1}**", key=f"page_{p}"): 
                        st.session_state.page = p
                        st.rerun()

        with col3:
            if st.session_state.page < total_pages - 1:
                if st.button("Next Page ➡️"):
                    st.session_state.page += 1
                    st.rerun()
        
        # Save feedback
        if st.button("Save My Responses"):
            with open(FEEDBACK_FILE, "w") as f:
                json.dump(st.session_state.feedback, f)
            st.success("Responses Saved!")

        # Download Excel
        def create_excel_download(df):
            wb = Workbook()
            ws = wb.active
            headers = df.columns.tolist()
            ws.append(headers + ["Review Status","Comments"])

            for idx, row in df.iterrows():
                project_id = str(row["Project Id"])
                feedback_data = st.session_state.feedback.get(project_id, {})
                status = st.session_state.feedback.get(project_id, {}).get("Quality", "Status Yet to be Updated")
                comment = feedback_data.get("comment", "")

                row_data = row.tolist() + [status,comment]
                ws.append(row_data)

                fill_color = {
                    "Correct": "00FF00",
                    "Incorrect": "FF0000"
                }.get(status, "FFFFFF")

                for col in range(1, len(headers) + 2):
                    ws.cell(row=idx + 2, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer

        if st.button("Download Excel"):
            excel_buffer = create_excel_download(filtered_df)
            st.download_button("Download the Excel file", excel_buffer, "qc_log.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
